"""
generate_quote_template.py -- 見積ブランクフォーマット生成スクリプト

使い方:
    python scripts/generate_quote_template.py --commodity マーケティング
    python scripts/generate_quote_template.py --all          # 全コモディティ
    python scripts/generate_quote_template.py --format csv   # CSV形式

オプション:
    --commodity  コモディティ大分類名（省略時は全生成）
    --all        全コモディティを生成
    --format     excel | csv  (default: excel)
    --outdir     出力ディレクトリ (default: output/)
    --rfq-id     RFQ番号（ヘッダ行に埋め込む例）
    --list       利用可能なコモディティ一覧を表示
"""
from __future__ import annotations

import argparse
import io
import sys
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd

PROJECT_ROOT = Path(__file__).parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from scripts.template_utils import (
    get_columns, get_benchmark_metrics, list_commodities,
    load_templates, STYLE,
)

# ---------------------------------------------------------------------------
# CSV生成
# ---------------------------------------------------------------------------
def generate_csv(commodity_group: str, rfq_id: str = "") -> str:
    """列ヘッダのみのブランクCSV文字列を返す"""
    cols = get_columns(commodity_group)
    col_names = [c["name"] for c in cols]
    # サンプル1行（空）
    df = pd.DataFrame(columns=col_names)
    # 1行目: label_ja の行を追加（コメント行として）
    labels = [c.get("label_ja", c["name"]) for c in cols]
    df_with_label = pd.concat([
        pd.DataFrame([labels], columns=col_names),
        df,
    ], ignore_index=True)
    return df_with_label.to_csv(index=False, encoding="utf-8-sig")


def save_csv(commodity_group: str, outdir: Path, rfq_id: str = "") -> Path:
    cfg = load_templates()
    name_en = cfg["commodities"][commodity_group].get("name_en", commodity_group)
    outdir.mkdir(parents=True, exist_ok=True)
    outpath = outdir / f"{name_en}_quote_template.csv"
    csv_str = generate_csv(commodity_group, rfq_id)
    outpath.write_text(csv_str, encoding="utf-8-sig")
    return outpath


# ---------------------------------------------------------------------------
# Excel生成
# ---------------------------------------------------------------------------
def generate_excel_bytes(
    commodity_group: str,
    rfq_id: str = "",
    rfq_name: str = "",
) -> bytes:
    """Excelバイト列を返す（Streamlit ダウンロード用）"""
    buf = io.BytesIO()
    _write_excel(commodity_group, buf, rfq_id=rfq_id, rfq_name=rfq_name)
    return buf.getvalue()


def save_excel(
    commodity_group: str,
    outdir: Path,
    rfq_id: str = "",
    rfq_name: str = "",
) -> Path:
    cfg = load_templates()
    name_en = cfg["commodities"][commodity_group].get("name_en", commodity_group)
    outdir.mkdir(parents=True, exist_ok=True)
    outpath = outdir / f"{name_en}_quote_template.xlsx"
    with open(outpath, "wb") as f:
        f.write(generate_excel_bytes(commodity_group, rfq_id, rfq_name))
    return outpath


def _write_excel(
    commodity_group: str,
    dest: Any,          # ファイルパス or BytesIO
    rfq_id: str = "",
    rfq_name: str = "",
) -> None:
    """openpyxl で Excel を書き出す（3シート構成）"""
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl が必要です: uv add openpyxl")

    cols = get_columns(commodity_group)
    benchmarks = get_benchmark_metrics(commodity_group)
    cfg = load_templates()
    commodity_cfg = cfg["commodities"][commodity_group]
    desc = commodity_cfg.get("description", "")

    wb = openpyxl.Workbook()

    # =====================================================================
    # Sheet 1: Instructions
    # =====================================================================
    ws_inst = wb.active
    ws_inst.title = "Instructions"

    ws_inst["A1"] = "見積ブランクフォーマット 記入要領"
    ws_inst["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws_inst["A2"] = f"コモディティ: {commodity_group}（{commodity_cfg.get('name_en', '')}）"
    ws_inst["A3"] = f"対象: {desc}"
    ws_inst["A4"] = f"生成日: {date.today()}"
    if rfq_id:
        ws_inst["A5"] = f"RFQ番号: {rfq_id}"
    if rfq_name:
        ws_inst["A6"] = f"案件名: {rfq_name}"

    ws_inst["A8"] = "【記入ルール】"
    ws_inst["A8"].font = Font(bold=True)
    rules = [
        "1. 黄色セル（必須）は必ず記入してください。",
        "2. 青色列はすべてのコモディティ共通のコア項目です。",
        "3. 緑色列はこのコモディティ固有の拡張項目です。",
        "4. line_amount = quantity × unit_price で計算してください。",
        "5. 通貨はJPY（円）を基本とします。外貨の場合は currency 列に記入。",
        "6. 日付は YYYY-MM-DD 形式で記入してください。",
        "7. フラグ列（*_flag）は 1=Yes / 0=No で記入してください。",
    ]
    for i, r in enumerate(rules):
        ws_inst[f"A{9+i}"] = r

    ws_inst["A17"] = "【ベンチマーク指標】"
    ws_inst["A17"].font = Font(bold=True)
    if benchmarks:
        ws_inst["A18"] = "指標名"
        ws_inst["B18"] = "計算式"
        ws_inst["C18"] = "単位"
        for i, bm in enumerate(benchmarks):
            ws_inst[f"A{19+i}"] = bm.get("metric_name", "")
            ws_inst[f"B{19+i}"] = bm.get("formula", "")
            ws_inst[f"C{19+i}"] = bm.get("display_unit", "")

    ws_inst.column_dimensions["A"].width = 50
    ws_inst.column_dimensions["B"].width = 35
    ws_inst.column_dimensions["C"].width = 20

    # =====================================================================
    # Sheet 2: Quotation_Header
    # =====================================================================
    ws_hdr = wb.create_sheet("Quotation_Header")
    header_cols = [
        "rfq_id", "rfq_name", "commodity_group", "supplier_id", "supplier_name",
        "quotation_id", "quotation_date", "valid_until", "currency", "total_amount",
        "payment_terms", "contract_term", "lead_time", "start_date", "end_date",
        "trade_terms", "remarks",
    ]
    core_col_names = {c["name"] for c in cols if c.get("is_core")}

    # ヘッダ行（列名）
    for ci, cn in enumerate(header_cols, start=1):
        cell = ws_hdr.cell(row=1, column=ci, value=cn)
        cell.fill = PatternFill("solid", fgColor=STYLE["header_fill"])
        cell.font = Font(bold=True, color=STYLE["header_font"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_hdr.column_dimensions[get_column_letter(ci)].width = 18

    # ラベル行（label_ja）
    col_label_map = {c["name"]: c.get("label_ja", c["name"]) for c in cols}
    for ci, cn in enumerate(header_cols, start=1):
        label = col_label_map.get(cn, cn)
        cell = ws_hdr.cell(row=2, column=ci, value=label)
        cell.fill = PatternFill("solid", fgColor=STYLE["core_fill"])
        cell.font = Font(italic=True, color="444444")
        cell.alignment = Alignment(horizontal="center")

    # サンプル入力行（3行目）
    defaults = {
        "rfq_id": rfq_id or "RFQ-202603-0001",
        "rfq_name": rfq_name or "（案件名を入力）",
        "commodity_group": commodity_group,
        "currency": "JPY",
        "quotation_date": date.today().strftime("%Y-%m-%d"),
    }
    for ci, cn in enumerate(header_cols, start=1):
        val = defaults.get(cn, "")
        cell = ws_hdr.cell(row=3, column=ci, value=val)
        cell.alignment = Alignment(horizontal="left")

    ws_hdr.row_dimensions[1].height = STYLE["header_height"]
    ws_hdr.row_dimensions[2].height = STYLE["row_height"]
    ws_hdr.freeze_panes = "A3"

    # =====================================================================
    # Sheet 3: Quotation_Line（見積明細）
    # =====================================================================
    ws_line = wb.create_sheet("Quotation_Line")

    req_names = {c["name"] for c in cols if c.get("required")}

    # 行1: 列名
    for ci, col in enumerate(cols, start=1):
        cell = ws_line.cell(row=1, column=ci, value=col["name"])
        if col.get("is_core"):
            bg = STYLE["core_fill"]
        else:
            bg = STYLE["ext_fill"]
        if col["name"] in req_names:
            bg = STYLE["req_fill"]
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color=STYLE["header_fill"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # 列幅
        width = STYLE["col_width_notes"] if col["name"] in ("specification", "remarks", "service_scope") \
                else STYLE["col_width_default"]
        ws_line.column_dimensions[get_column_letter(ci)].width = width

    # 行2: 日本語ラベル
    for ci, col in enumerate(cols, start=1):
        cell = ws_line.cell(row=2, column=ci, value=col.get("label_ja", col["name"]))
        cell.font = Font(italic=True, color="444444", size=9)
        cell.alignment = Alignment(horizontal="center")

    # 行3: 必須/任意 表示
    for ci, col in enumerate(cols, start=1):
        req_label = "【必須】" if col.get("required") else ""
        cell = ws_line.cell(row=3, column=ci, value=req_label)
        cell.font = Font(color="C00000", bold=True, size=8)
        cell.alignment = Alignment(horizontal="center")

    # 行4-13: 入力行（空白）
    for r in range(4, 14):
        for ci, col in enumerate(cols, start=1):
            ws_line.cell(row=r, column=ci, value="")
        ws_line.row_dimensions[r].height = STYLE["row_height"]

    ws_line.row_dimensions[1].height = STYLE["header_height"]
    ws_line.row_dimensions[2].height = STYLE["row_height"]
    ws_line.freeze_panes = "A4"  # 3行ヘッダを固定

    wb.save(dest)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="見積ブランクフォーマット生成")
    parser.add_argument("--commodity", default=None, help="コモディティ大分類名")
    parser.add_argument("--all",    action="store_true", help="全コモディティを生成")
    parser.add_argument("--format", default="excel", choices=["excel", "csv"])
    parser.add_argument("--outdir", default=None, help="出力ディレクトリ")
    parser.add_argument("--rfq-id", default="", help="RFQ番号（ヘッダ埋め込み）")
    parser.add_argument("--list",   action="store_true", help="コモディティ一覧表示")
    args = parser.parse_args()

    if args.list:
        print("利用可能なコモディティ:")
        for c in list_commodities():
            print(f"  - {c}")
        sys.exit(0)

    outdir = Path(args.outdir) if args.outdir else PROJECT_ROOT / "output"
    targets = list_commodities() if args.all else ([args.commodity] if args.commodity else list_commodities())

    for cg in targets:
        try:
            if args.format == "excel":
                path = save_excel(cg, outdir, rfq_id=args.rfq_id)
            else:
                path = save_csv(cg, outdir, rfq_id=args.rfq_id)
            print(f"[OK] {cg} -> {path}")
        except Exception as e:
            print(f"[ERR] {cg}: {e}")
